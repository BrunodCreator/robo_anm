from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# import preencher_formulario
from time import sleep

    
def capturar_todos_os_dados(func_navegador):
    # Captura os dados da primeira tabela
    elemento = func_navegador.find_element(By.CSS_SELECTOR, '#ctl00_ContentPlaceHolder1_dvResultado > table.tabelaFormulario')
    texto = elemento.text

    # Transforma os dados da primeira tabela em um dicionário
    dados_gerais = {}
    for linha in texto.split("\n"):
        if " : " in linha:
            chave, valor = linha.split(" : ", 1)
            dados_gerais[chave.strip()] = valor.strip()

    print("Dados gerais capturados:", dados_gerais)

    # Captura os dados da segunda tabela
    tabela = func_navegador.find_element(By.CSS_SELECTOR, '#ctl00_ContentPlaceHolder1_dvResultado > table.tabelaRelatorio')
    linhas = tabela.find_elements(By.CSS_SELECTOR, "tbody > tr")

    # Combina os dados das duas tabelas em uma lista de dicionários
    dados_completos = []
    for linha in linhas:
        colunas = linha.find_elements(By.TAG_NAME, "td")
        if len(colunas) >= 6:
            # Cria um dicionário com os dados da linha
            linha_dados = {
                "Arrecadador (Empresa)": colunas[1].text,
                "Qtde Títulos": colunas[2].text,
                "Operação": colunas[3].text,
                "Recolhimento CFEM": colunas[4].text,
                "% Recolhimento CFEM": colunas[5].text
                
            }

            # Adiciona os dados gerais (primeira tabela) a cada linha
            linha_dados.update(dados_gerais)
            dados_completos.append(linha_dados)
        else:
            print(f"Linha ignorada por não ter colunas suficientes: {linha.text}")

    # Exibe os dados completos
    for dado in dados_completos:
        print(dado)

    return dados_completos


def salvar_dados_completos_planilha(dados_completos, nome_arquivo=".xlsx"):
    
    
    # Caminho para salvar o arquivo na área de trabalho
    caminho_arquivo = os.path.join(os.path.expanduser("~"), "Desktop", nome_arquivo)

    # Verifica se a planilha já existe
    if os.path.exists(caminho_arquivo):
        # Abre a planilha existente
        workbook = load_workbook(caminho_arquivo)
    else:
        # Cria uma nova planilha
        workbook = Workbook()

    # Seleciona a primeira folha da planilha
    folha = workbook.active
    folha.title = "Arrecadadores"

    # Adiciona o cabeçalho (apenas se a planilha for nova)
    if folha.max_row == 1:
        folha.append([
            "Arrecadador (Empresa)", "Qtde Títulos", "Operação",
            "Recolhimento CFEM", "% Recolhimento CFEM",  # Dados da segunda tabela
            "Ano", "Arrecadação por", "Ordenação por", "Substância Agrupadora", "Substância", "Região", "Estado", "Municipio" # Dados gerais (primeira tabela)
        ])

    # Adiciona os dados capturados
    for dado in dados_completos:
        folha.append([
            dado["Arrecadador (Empresa)"],
            dado["Qtde Títulos"],
            dado["Operação"],
            dado["Recolhimento CFEM"],
            dado["% Recolhimento CFEM"],
            dado.get("Ano", ""),  # Dados gerais (adicionais)
            dado.get("Arrecadação por", ""),
            dado.get("Ordenação por", ""),
            dado.get("Substância Agrupadora", ""),
            dado.get("Substância", ""),
            dado.get("Região", ""),
            dado.get("Estado", ""),
            dado.get("Municipio", "")
        ])

    # Salva a planilha
    workbook.save(caminho_arquivo)
    print(f"Dados salvos na planilha '{caminho_arquivo}' com sucesso!")
    
    
    
# def mesclar_planilhas(nome_arquivo="resultado_final.xlsx"):
#     """Lê todos os arquivos gerados pelos processos e mescla em um único arquivo Excel."""

#     # Encontrar todos os arquivos criados pelos processos
#     arquivos = [f for f in os.listdir(os.path.expanduser("~\\Desktop")) if f.startswith(f"{nome_arquivo}") and f.endswith(".xlsx")]

#     if not arquivos:
#         print("Nenhum arquivo de processo encontrado para mesclar.")
#         return

#     # Criar um novo workbook consolidado
#     caminho_final = os.path.join(os.path.expanduser("~"), "Desktop", nome_arquivo)
#     workbook_final = Workbook()
#     folha_final = workbook_final.active
#     folha_final.title = "Dados Consolidados"

#     # Abrir o primeiro arquivo e copiar o cabeçalho
#     primeiro_arquivo = os.path.join(os.path.expanduser("~"), "Desktop", arquivos[0])
#     workbook = load_workbook(primeiro_arquivo)
#     folha = workbook.active
#     cabecalho = [cell.value for cell in folha[1]]  # Copiar cabeçalho da primeira linha
#     folha_final.append(cabecalho)

#     # Adicionar os dados de todos os arquivos
#     for arquivo in arquivos:
#         caminho_arquivo = os.path.join(os.path.expanduser("~"), "Desktop", arquivo)
#         workbook = load_workbook(caminho_arquivo)
#         folha = workbook.active

#         for linha in folha.iter_rows(min_row=2, values_only=True):  # Pular cabeçalho
#             folha_final.append(linha)

#     # Salvar o arquivo consolidado
#     workbook_final.save(caminho_final)
#     print(f"Dados mesclados no arquivo '{caminho_final}' com sucesso!")

#     # # Opcional: Apagar arquivos individuais dos processos após a fusão
#     # for arquivo in arquivos:
#     #     os.remove(os.path.join(os.path.expanduser("~"), "Desktop", arquivo))
#     #     print(f"Arquivo {arquivo} removido após a fusão.")
